import requests
import time
import json
import csv
from datetime import datetime
import threading
from queue import Queue, Empty
import os
from werkzeug.utils import secure_filename
import logging
from typing import Dict, List, Optional
import re
from collections import defaultdict
import uuid
from io import BytesIO
import zipfile
from flask import Flask, send_file, request, jsonify, render_template
from flask_socketio import SocketIO, emit
from config import Config
from datetime import timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, ListFlowable, ListItem
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

app = Flask(__name__)
app.config.from_object(Config)
# Restrict CORS origins for Socket.IO to local development origins (use configured port)
_port = getattr(Config, 'PORT', 8000)
socketio = SocketIO(
    app,
    cors_allowed_origins=[
        f"http://localhost:{_port}",
        f"http://127.0.0.1:{_port}",
    ],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('breach_scanner.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class BreachChecker:
    """HaveIBeenPwned API integration with enhanced error handling"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://haveibeenpwned.com/api/v3"
        self.session = requests.Session()
        self.session.headers.update({
            'hibp-api-key': api_key,
            'User-Agent': 'NWCRC-Email-Check/1.0 (+https://github.com/Yenn503/NWCRC-Email-Check)',
            'Accept': 'application/json'
        })
        
    def _get_with_retry(self, url: str, timeout: int = 30):
        """Perform GET with a single retry on 429 honoring Retry-After."""
        response = self.session.get(url, timeout=timeout)
        if response.status_code == 429:
            retry_after = response.headers.get('Retry-After')
            try:
                delay = int(retry_after) if retry_after else 60
                delay = max(1, min(delay, 60))
            except Exception:
                delay = 60
            logger.warning(f"Rate limited. Retrying after {delay}s for URL {url}")
            time.sleep(delay)
            return self.session.get(url, timeout=timeout)
        return response

    def check_breaches(self, email: str) -> Dict:
        """Check if email appears in any breaches"""
        try:
            url = f"{self.base_url}/breachedaccount/{email}"
            response = self._get_with_retry(url, timeout=30)
            
            if response.status_code == 200:
                breaches = response.json()
                return {
                    'status': 'compromised' if breaches else 'clean',
                    'breaches': breaches or [],
                    'breach_count': len(breaches) if breaches else 0,
                    'severity': self._calculate_severity(breaches) if breaches else 'clean'
                }
            elif response.status_code == 404:
                return {
                    'status': 'clean',
                    'breaches': [],
                    'breach_count': 0,
                    'severity': 'clean'
                }
            elif response.status_code == 429:
                retry_after = response.headers.get('Retry-After')
                logger.warning(f"Rate limit exceeded for {email}; Retry-After={retry_after}")
                return {
                    'status': 'error',
                    'error': 'Rate limit exceeded',
                    'breaches': [],
                    'breach_count': 0,
                    'retry_after': retry_after
                }
            elif response.status_code == 401:
                logger.error("Unauthorized (401) from HIBP. Check that your API key is valid and active.")
                return {
                    'status': 'error',
                    'error': 'Unauthorized: Invalid or inactive API key',
                    'breaches': [],
                    'breach_count': 0
                }
            else:
                logger.error(f"API error for {email}: {response.status_code}")
                return {
                    'status': 'error',
                    'error': f'API error: {response.status_code}',
                    'breaches': [],
                    'breach_count': 0
                }
                
        except requests.exceptions.RequestException as e:
            logger.error(f"Network error checking {email}: {str(e)}")
            return {
                'status': 'error',
                'error': f'Network error: {str(e)}',
                'breaches': [],
                'breach_count': 0
            }
    
    def check_pastes(self, email: str) -> Dict:
        """Check if email appears in any pastes"""
        try:
            url = f"{self.base_url}/pasteaccount/{email}"
            response = self._get_with_retry(url, timeout=30)
            
            if response.status_code == 200:
                pastes = response.json()
                return {
                    'pastes': pastes or [],
                    'paste_count': len(pastes) if pastes else 0
                }
            elif response.status_code == 404:
                return {
                    'pastes': [],
                    'paste_count': 0
                }
            else:
                return {
                    'pastes': [],
                    'paste_count': 0
                }
                
        except requests.exceptions.RequestException:
            return {
                'pastes': [],
                'paste_count': 0
            }
    
    def _calculate_severity(self, breaches: List[Dict]) -> str:
        """Calculate severity based on breach characteristics"""
        if not breaches:
            return 'clean'
        
        high_severity_indicators = ['passwords', 'credit cards', 'social security numbers']
        verified_breaches = sum(1 for b in breaches if b.get('IsVerified', False))
        sensitive_breaches = sum(1 for b in breaches if b.get('IsSensitive', False))
        
        # Check for high-risk data classes
        high_risk_count = 0
        for breach in breaches:
            data_classes = [dc.lower() for dc in breach.get('DataClasses', [])]
            if any(indicator in ' '.join(data_classes) for indicator in high_severity_indicators):
                high_risk_count += 1
        
        if high_risk_count > 0 or sensitive_breaches > 2:
            return 'critical'
        elif verified_breaches > 3 or len(breaches) > 5:
            return 'high'
        elif len(breaches) > 2:
            return 'medium'
        else:
            return 'low'

class BatchProcessor:
    """Enhanced batch processing with real-time updates and control"""
    
    def __init__(self, breach_checker: BreachChecker):
        self.breach_checker = breach_checker
        self.scan_queue = Queue()
        self.scan_results = []
        self.scan_progress = {
            'total': 0,
            'completed': 0,
            'current_email': '',
            'status': 'idle',
            'start_time': None,
            'estimated_completion': None
        }
        self.current_batch_id = None
        self.is_processing = False
        self.is_paused = False
        self.should_stop = False
        self.processing_thread = None
        self.rate_limit_delay = 60 / Config.RATE_LIMIT_PER_MINUTE  # Convert to seconds between requests
        self._lock = threading.Lock()
        
    def start_batch_scan(self, emails: List[str]) -> str:
        """Start a new batch scan"""
        if self.is_processing:
            raise ValueError("Another scan is already in progress")
        
        # Validate and clean emails
        valid_emails = []
        for email in emails:
            email = email.strip().lower()
            if self._is_valid_email(email):
                valid_emails.append(email)
        
        if not valid_emails:
            raise ValueError("No valid emails provided")
        
        # Remove duplicates while preserving order
        seen = set()
        unique_emails = []
        for email in valid_emails:
            if email not in seen:
                seen.add(email)
                unique_emails.append(email)
        
        # Initialize batch
        self.current_batch_id = str(uuid.uuid4())[:8]
        with self._lock:
            self.scan_results = []
        self.scan_queue = Queue()
        
        # Add emails to queue
        for email in unique_emails:
            self.scan_queue.put(email)
        
        # Initialize progress
        with self._lock:
            self.scan_progress = {
                'total': len(unique_emails),
                'completed': 0,
                'current_email': '',
                'status': 'running',
                'start_time': datetime.now().isoformat(),
                'estimated_completion': None,
                'batch_id': self.current_batch_id
            }
        
        # Start processing thread
        self.is_processing = True
        self.is_paused = False
        self.should_stop = False
        self.processing_thread = threading.Thread(target=self._process_batch)
        self.processing_thread.daemon = True
        self.processing_thread.start()
        
        logger.info(f"Started batch scan {self.current_batch_id} with {len(unique_emails)} emails")
        return self.current_batch_id
    
    def pause_batch(self):
        """Pause the current batch processing"""
        if self.is_processing:
            self.is_paused = True
            with self._lock:
                self.scan_progress['status'] = 'paused'
            logger.info(f"Paused batch scan {self.current_batch_id}")
    
    def resume_batch(self):
        """Resume the paused batch processing"""
        if self.is_processing and self.is_paused:
            self.is_paused = False
            with self._lock:
                self.scan_progress['status'] = 'running'
            logger.info(f"Resumed batch scan {self.current_batch_id}")
    
    def stop_batch(self):
        """Stop the current batch processing"""
        if self.is_processing:
            self.should_stop = True
            with self._lock:
                self.scan_progress['status'] = 'stopped'
            # Try to join the processing thread for clean stop
            if self.processing_thread and self.processing_thread.is_alive():
                self.processing_thread.join(timeout=1)
            logger.info(f"Stopped batch scan {self.current_batch_id}")
    
    def _process_batch(self):
        """Process emails in the batch queue"""
        try:
            while not self.scan_queue.empty() and not self.should_stop:
                # Handle pause
                while self.is_paused and not self.should_stop:
                    time.sleep(0.5)
                
                if self.should_stop:
                    break
                
                try:
                    email = self.scan_queue.get_nowait()
                    with self._lock:
                        self.scan_progress['current_email'] = email
                    
                    # Emit progress update
                    socketio.emit('scan_progress', self.scan_progress)
                    
                    # Check email
                    result = self._scan_single_email(email)
                    with self._lock:
                        self.scan_results.append(result)
                    
                    # Update progress
                    with self._lock:
                        self.scan_progress['completed'] += 1
                        self._update_estimated_completion()
                    
                    # Emit result
                    socketio.emit('scan_result', result)
                    socketio.emit('scan_progress', self.get_status_snapshot())
                    
                    # Rate limiting
                    if not self.scan_queue.empty():
                        time.sleep(self.rate_limit_delay)
                    
                except Empty:
                    break
                except Exception as e:
                    logger.error(f"Error processing email {email}: {str(e)}")
                    continue
            
            # Batch completed
            with self._lock:
                self.scan_progress['status'] = 'completed' if not self.should_stop else 'stopped'
                self.scan_progress['current_email'] = ''
                self.is_processing = False
            
            # Emit final update
            socketio.emit('scan_progress', self.get_status_snapshot())
            socketio.emit('batch_complete', {
                'batch_id': self.current_batch_id,
                'total_results': len(self.get_results_snapshot()),
                'statistics': self.get_batch_statistics()
            })
            
            logger.info(f"Batch scan {self.current_batch_id} completed with {len(self.scan_results)} results")
            
        except Exception as e:
            logger.error(f"Batch processing error: {str(e)}")
            with self._lock:
                self.scan_progress['status'] = 'error'
                self.is_processing = False
            socketio.emit('scan_error', {'error': str(e)})
    
    def _scan_single_email(self, email: str) -> Dict:
        """Scan a single email and return comprehensive results"""
        timestamp = datetime.now().isoformat()
        
        try:
            # Check breaches
            breach_result = self.breach_checker.check_breaches(email)
            
            # Check pastes (optional, can be disabled to improve speed)
            paste_result = {'pastes': [], 'paste_count': 0}
            if Config.CHECK_PASTES:
                paste_result = self.breach_checker.check_pastes(email)
            
            # Combine results
            result = {
                'email': email,
                'timestamp': timestamp,
                'status': breach_result['status'],
                'breaches': breach_result.get('breaches', []),
                'breach_count': breach_result.get('breach_count', 0),
                'severity': breach_result.get('severity', 'unknown'),
                'pastes': paste_result.get('pastes', []),
                'paste_count': paste_result.get('paste_count', 0),
                'error': breach_result.get('error')
            }
            
            return result
            
        except Exception as e:
            logger.error(f"Error scanning {email}: {str(e)}")
            return {
                'email': email,
                'timestamp': timestamp,
                'status': 'error',
                'error': str(e),
                'breaches': [],
                'breach_count': 0,
                'pastes': [],
                'paste_count': 0
            }
    
    def _update_estimated_completion(self):
        """Update estimated completion time"""
        if self.scan_progress['completed'] > 0 and self.scan_progress['start_time']:
            # Parse start_time from ISO string if needed
            start_time = self.scan_progress['start_time']
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time)
            elapsed = (datetime.now() - start_time).total_seconds()
            rate = self.scan_progress['completed'] / elapsed if elapsed > 0 else 0
            remaining = self.scan_progress['total'] - self.scan_progress['completed']
            if rate > 0:
                eta_seconds = remaining / rate
                eta = datetime.now() + timedelta(seconds=eta_seconds)
                self.scan_progress['estimated_completion'] = eta.isoformat()
    
    def _is_valid_email(self, email: str) -> bool:
        """Validate email format"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    def get_batch_statistics(self) -> Dict:
        """Get comprehensive batch statistics"""
        with self._lock:
            if not self.scan_results:
                return {}
            results_copy = list(self.scan_results)
            progress_copy = dict(self.scan_progress)
        
        stats = {
            'total_emails': len(results_copy),
            'clean_emails': len([r for r in results_copy if r['status'] == 'clean']),
            'compromised_emails': len([r for r in results_copy if r['status'] == 'compromised']),
            'error_emails': len([r for r in results_copy if r['status'] == 'error']),
            'total_breaches': sum(r.get('breach_count', 0) for r in results_copy),
            'total_pastes': sum(r.get('paste_count', 0) for r in results_copy),
        }
        
        # Calculate processing time
        if progress_copy.get('start_time'):
            start_time = progress_copy['start_time']
            if isinstance(start_time, str):
                try:
                    start_time = datetime.fromisoformat(start_time)
                except Exception:
                    start_time = None
            if start_time:
                processing_time = (datetime.now() - start_time).total_seconds()
                stats['processing_time'] = processing_time
        
        # Severity breakdown
        severity_counts = defaultdict(int)
        for result in results_copy:
            if result['status'] == 'compromised':
                severity_counts[result.get('severity', 'unknown')] += 1
        stats['severity_breakdown'] = dict(severity_counts)
        
        # Top breaches
        breach_counts = defaultdict(int)
        for result in results_copy:
            for breach in result.get('breaches', []):
                breach_counts[breach.get('Name', 'Unknown')] += 1
        stats['top_breaches'] = dict(sorted(breach_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        
        return stats

    def get_status_snapshot(self) -> Dict:
        with self._lock:
            return dict(self.scan_progress)

    def get_results_snapshot(self) -> List[Dict]:
        with self._lock:
            return list(self.scan_results)

try:
    Config.validate_config()
    breach_checker = BreachChecker(Config.HIBP_API_KEY)
    batch_processor = BatchProcessor(breach_checker)
    logger.info("Application initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize application: {str(e)}")
    raise

@app.route("/")
def index():
    return render_template("index.html")

@app.route('/scan-single', methods=['POST'])
def scan_single():
    """Scan a single email address"""
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    
    if not email:
        return jsonify({'error': 'Email address is required'}), 400
    
    if not batch_processor._is_valid_email(email):
        return jsonify({'error': 'Invalid email format'}), 400
    
    try:
        result = batch_processor._scan_single_email(email)
        logger.info(f"Single scan completed for {email}: {result['status']}")
        return jsonify(result)
    except Exception as e:
        logger.error(f"Single scan failed for {email}: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/scan-batch', methods=['POST'])
def scan_batch():
    """Start a batch scan"""
    data = request.get_json()
    emails = data.get('emails', [])
    
    if not emails:
        return jsonify({'error': 'Email list is required'}), 400
    
    # Handle both string (newline-separated) and array inputs
    if isinstance(emails, str):
        emails = [email.strip() for email in emails.split('\n') if email.strip()]
    
    try:
        batch_id = batch_processor.start_batch_scan(emails)
        return jsonify({
            'message': 'Batch scan started',
            'batch_id': batch_id,
            'total_emails': batch_processor.scan_progress['total']
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Batch scan failed: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/batch-control', methods=['POST'])
def batch_control():
    """Control batch processing (pause/resume/stop)"""
    data = request.get_json()
    action = data.get('action')
    
    if not batch_processor.is_processing:
        return jsonify({'error': 'No active batch scan'}), 400
    
    try:
        if action == 'pause':
            batch_processor.pause_batch()
        elif action == 'resume':
            batch_processor.resume_batch()
        elif action == 'stop':
            batch_processor.stop_batch()
        else:
            return jsonify({'error': 'Invalid action'}), 400
        
        return jsonify({'message': f'Batch scan {action}ed successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/batch-status')
def batch_status():
    """Get current batch status"""
    return jsonify({
        'progress': batch_processor.get_status_snapshot(),
        'results_count': len(batch_processor.get_results_snapshot()),
        'is_processing': batch_processor.is_processing
    })

@app.route('/batch-results')
def batch_results():
    """Get current batch results"""
    return jsonify({
        'results': batch_processor.get_results_snapshot(),
        'statistics': batch_processor.get_batch_statistics(),
        'progress': batch_processor.get_status_snapshot()
    })

@app.route('/export-results', methods=['POST'])
def export_results():
    """Enhanced export scan results to multiple file formats"""
    data = request.get_json()
    format_type = data.get('format', 'json')
    results = data.get('results', batch_processor.scan_results)
    export_options = data.get('options', {})
    
    if not results:
        return jsonify({'error': 'No results to export'}), 400
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    batch_id = batch_processor.current_batch_id or 'unknown'
    
    try:
        if format_type == 'json':
            return export_json(results, batch_id, timestamp, export_options)
        elif format_type == 'csv':
            return export_csv(results, batch_id, timestamp, export_options)
        elif format_type in ['excel', 'xlsx']:
            if not EXCEL_AVAILABLE:
                return jsonify({'error': 'Excel export not available - openpyxl not installed'}), 400
            return export_excel(results, batch_id, timestamp, export_options)
        elif format_type == 'pdf' and PDF_AVAILABLE:
            return export_pdf(results, batch_id, timestamp, export_options)
        elif format_type == 'zip':
            return export_zip_package(results, batch_id, timestamp, export_options)
        else:
            available_formats = ['json', 'csv']
            if EXCEL_AVAILABLE:
                available_formats.append('excel')
            if PDF_AVAILABLE:
                available_formats.append('pdf')
            available_formats.append('zip')
            return jsonify({
                'error': f'Invalid or unavailable format type. Available formats: {", ".join(available_formats)}'
            }), 400
        
    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

def write_json_file(results, batch_id, timestamp, options):
    """Export results as JSON with enhanced metadata"""
    filename = f"breach_scan_results_{batch_id}_{timestamp}.json"
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    # Enhanced export data with comprehensive metadata
    export_data = {
        'metadata': {
            'batch_id': batch_id,
            'export_timestamp': datetime.now().isoformat(),
            'export_format': 'json',
            'total_emails': len(results),
            'scanner_version': '1.0',
            'api_version': 'v3',
            'export_options': options
        },
        'statistics': batch_processor.get_batch_statistics(),
        'summary': {
            'clean_emails': len([r for r in results if r['status'] == 'clean']),
            'compromised_emails': len([r for r in results if r['status'] in ['compromised', 'breached']]),
            'error_emails': len([r for r in results if r['status'] == 'error']),
            'total_breaches': sum(r.get('breach_count', 0) for r in results),
            'total_pastes': sum(r.get('paste_count', 0) for r in results)
        },
        'results': results
    }
    
    # Apply export options
    if options.get('exclude_clean', False):
        export_data['results'] = [r for r in results if r['status'] != 'clean']
    
    if options.get('only_high_severity', False):
        export_data['results'] = [r for r in export_data['results'] 
                                if r.get('severity') in ['high', 'critical']]
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2)
    logger.info(f"JSON results exported to {filename}")
    return filename

def export_json(results, batch_id, timestamp, options):
    filename = write_json_file(results, batch_id, timestamp, options)
    return jsonify({'message': 'Results exported successfully', 'filename': filename, 'format': 'json'})

def write_csv_file(results, batch_id, timestamp, options):
    """Export results as CSV with customizable columns"""
    filename = f"breach_scan_results_{batch_id}_{timestamp}.csv"
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    # Apply filters
    filtered_results = results
    if options.get('exclude_clean', False):
        filtered_results = [r for r in filtered_results if r['status'] != 'clean']
    
    if options.get('only_high_severity', False):
        filtered_results = [r for r in filtered_results if r.get('severity') in ['high', 'critical']]
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # Enhanced CSV headers
        headers = [
            'Email', 'Status', 'Severity', 'Breach Count', 'Paste Count', 
            'Breaches', 'Pastes', 'Data Classes', 'Verified Breaches', 
            'Sensitive Breaches', 'Timestamp'
        ]
        writer.writerow(headers)
        
        for result in filtered_results:
            breaches_str = '; '.join([
                f"{b.get('Name', 'Unknown')} ({b.get('BreachDate', '')})" 
                for b in result.get('breaches', [])
            ])
            pastes_str = '; '.join([
                f"{p.get('Source', '')} ({p.get('Date', '')})" 
                for p in result.get('pastes', [])
            ])
            # Extract data classes
            data_classes = set()
            for breach in result.get('breaches', []):
                data_classes.update(breach.get('DataClasses', []))
            data_classes_str = '; '.join(sorted(data_classes))
            # Count verified and sensitive breaches
            verified_breaches = sum(1 for b in result.get('breaches', []) if b.get('IsVerified', False))
            sensitive_breaches = sum(1 for b in result.get('breaches', []) if b.get('IsSensitive', False))
            writer.writerow([
                result['email'],
                result['status'],
                result.get('severity', 'unknown'),
                result.get('breach_count', 0),
                result.get('paste_count', 0),
                breaches_str,
                pastes_str,
                data_classes_str,
                verified_breaches,
                sensitive_breaches,
                result['timestamp']
            ])
    logger.info(f"CSV results exported to {filename}")
    return filename

def export_csv(results, batch_id, timestamp, options):
    filename = write_csv_file(results, batch_id, timestamp, options)
    return jsonify({'message': 'Results exported successfully', 'filename': filename, 'format': 'csv'})

def write_excel_file(results, batch_id, timestamp, options):
    """Export results as Excel with formatting and multiple sheets"""
    if not EXCEL_AVAILABLE:
        raise RuntimeError('Excel export not available - openpyxl not installed')
    
    filename = f"breach_scan_results_{batch_id}_{timestamp}.xlsx"
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    # Apply filters
    filtered_results = results
    if options.get('exclude_clean', False):
        filtered_results = [r for r in filtered_results if r['status'] != 'clean']
    
    if options.get('only_high_severity', False):
        filtered_results = [r for r in filtered_results if r.get('severity') in ['high', 'critical']]
    
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Create Summary sheet
    summary_sheet = workbook.create_sheet("Summary")
    stats = batch_processor.get_batch_statistics()
    
    # Summary data
    summary_data = [
        ["Batch ID", batch_id],
        ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Emails Processed", len(results)],
        ["Clean Emails", stats.get('clean_emails', 0)],
        ["Compromised Emails", stats.get('compromised_emails', 0)],
        ["Error Emails", stats.get('error_emails', 0)],
        ["Total Breaches Found", stats.get('total_breaches', 0)],
        ["Total Pastes Found", stats.get('total_pastes', 0)],
        ["Processing Time", f"{stats.get('processing_time', 0):.1f} seconds" if stats.get('processing_time') else 'N/A']
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        summary_sheet.cell(row=row_idx, column=2, value=value)
    
    # Create Results sheet
    results_sheet = workbook.create_sheet("Results")
    
    # Headers
    headers = [
        'Email', 'Status', 'Severity', 'Breach Count', 'Paste Count', 
        'Breaches', 'Data Classes', 'Timestamp'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = results_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data rows
    for row_idx, result in enumerate(filtered_results, 2):
        results_sheet.cell(row=row_idx, column=1, value=result['email'])
        status_cell = results_sheet.cell(row=row_idx, column=2, value=result['status'])
        if result['status'] == 'compromised':
            status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif result['status'] == 'clean':
            status_cell.fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
        elif result['status'] == 'error':
            status_cell.fill = PatternFill(start_color="FFD43B", end_color="FFD43B", fill_type="solid")
        results_sheet.cell(row=row_idx, column=3, value=result.get('severity', 'unknown'))
        results_sheet.cell(row=row_idx, column=4, value=result.get('breach_count', 0))
        results_sheet.cell(row=row_idx, column=5, value=result.get('paste_count', 0))
        breaches_str = '; '.join([b.get('Name', 'Unknown') for b in result.get('breaches', [])])
        results_sheet.cell(row=row_idx, column=6, value=breaches_str)
        data_classes = set()
        for breach in result.get('breaches', []):
            data_classes.update(breach.get('DataClasses', []))
        results_sheet.cell(row=row_idx, column=7, value='; '.join(sorted(data_classes)))
        results_sheet.cell(row=row_idx, column=8, value=result['timestamp'])
    
    # Auto-adjust column widths
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(filepath)
    logger.info(f"Excel results exported to {filename}")
    return filename

def export_excel(results, batch_id, timestamp, options):
    try:
        filename = write_excel_file(results, batch_id, timestamp, options)
        return jsonify({'message': 'Results exported successfully', 'filename': filename, 'format': 'excel'})
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 400

def write_pdf_file(results, batch_id, timestamp, options):
    """Export results as a professional PDF report with branding and logo"""
    if not PDF_AVAILABLE:
        raise RuntimeError('PDF export not available - reportlab not installed')
    
    filename = f"breach_scan_report_{batch_id}_{timestamp}.pdf"
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    # Apply filters
    filtered_results = results
    if options.get('exclude_clean', False):
        filtered_results = [r for r in filtered_results if r['status'] != 'clean']
    
    if options.get('only_high_severity', False):
        filtered_results = [r for r in filtered_results if r.get('severity') in ['high', 'critical']]
    # Branding
    brand_primary = colors.HexColor('#0F4C81')  # Deep blue
    brand_accent = colors.HexColor('#51CF66')   # Green accent
    brand_light = colors.HexColor('#F5F7FA')    # Light background

    # Resolve logo path (fallback if primary not found)
    # Prefer absolute paths under the Flask app root
    base_static = os.path.join(app.root_path, 'static', 'img') if hasattr(app, 'root_path') else os.path.join('static', 'img')
    logo_path_candidates = [
        os.path.join(base_static, 'Logo.png'),
        os.path.join(base_static, 'Logo2.png'),
        os.path.join('static', 'img', 'Logo.png'),
        os.path.join('static', 'img', 'Logo2.png')
    ]
    logo_path = next((p for p in logo_path_candidates if os.path.exists(p)), None)

    # Document setup with margins
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=72,
        bottomMargin=54
    )
    styles = getSampleStyleSheet()
    story = []

    # Header/Footer drawing
    def draw_header_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        width, height = A4
        header_y = height - 50

        # Header separator line
        canvas_obj.setStrokeColor(brand_primary)
        canvas_obj.setFillColor(brand_primary)
        canvas_obj.setLineWidth(1)
        canvas_obj.line(36, header_y - 18, width - 36, header_y - 18)

        # Logo
        if logo_path:
            try:
                logo_h = 28
                canvas_obj.drawImage(logo_path, 36, header_y - logo_h, height=logo_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

        # Title and date
        canvas_obj.setFont('Helvetica-Bold', 12)
        canvas_obj.setFillColor(brand_primary)
        canvas_obj.drawString(200, header_y - 4, 'Email Breach Scan Report')
        canvas_obj.setFont('Helvetica', 9)
        canvas_obj.setFillColor(colors.black)
        canvas_obj.drawRightString(width - 36, header_y - 4, datetime.now().strftime('%Y-%m-%d %H:%M'))

        # Footer line and page number
        footer_y = 40
        canvas_obj.setStrokeColor(brand_primary)
        canvas_obj.line(36, footer_y + 8, width - 36, footer_y + 8)
        canvas_obj.setFont('Helvetica', 9)
        canvas_obj.setFillColor(colors.black)
        canvas_obj.drawString(36, footer_y, 'NWCRC Breach Scanner')
        canvas_obj.drawRightString(width - 36, footer_y, f'Page {canvas_obj.getPageNumber()}')
        canvas_obj.restoreState()

    # Cover title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=brand_primary,
        spaceAfter=16,
        alignment=1
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.black,
        alignment=1,
        spaceAfter=24
    )
    # Cover page content
    story.append(Spacer(1, 100))
    story.append(Paragraph('Email Breach Scan Report', title_style))
    story.append(Paragraph(f'Batch ID: {batch_id} • Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', subtitle_style))
    story.append(Spacer(1, 30))
    intro = Paragraph('This report summarizes the results of your email breach scan and includes a copy-friendly appendix of breached accounts.', styles['Normal'])
    story.append(intro)
    story.append(PageBreak())

    # Summary section
    stats = batch_processor.get_batch_statistics()
    summary_data = [
        ['Total Emails Scanned', str(len(results))],
        ['Clean Emails', str(stats.get('clean_emails', 0))],
        ['Compromised Emails', str(stats.get('compromised_emails', 0))],
        ['Error Emails', str(stats.get('error_emails', 0))],
        ['Total Breaches', str(stats.get('total_breaches', 0))],
        ['Processing Time', f"{stats.get('processing_time', 0):.1f}s" if stats.get('processing_time') else 'N/A']
    ]

    summary_table = Table(summary_data, colWidths=[2.6*inch, 1.6*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), brand_light),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 0.5, brand_primary),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
    ]))

    story.append(Paragraph('Scan Summary', styles['Heading2']))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    # Top breaches section (if available)
    top_breaches = stats.get('top_breaches') or {}
    if top_breaches:
        tb_rows = [['Breach Name', 'Occurrences']]
        tb_rows.extend([[name, str(count)] for name, count in top_breaches.items()])
        tb_table = Table(tb_rows, colWidths=[3.2*inch, 1.0*inch])
        tb_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 0.5, brand_primary),
        ]))
        story.append(Paragraph('Top Breaches', styles['Heading3']))
        story.append(tb_table)
        story.append(Spacer(1, 16))

    # Results section (limit to first 50 for readability). Breaches column removed.
    story.append(Paragraph('Detailed Results (First 50 entries)', styles['Heading2']))

    results_data = [['Email', 'Status', 'Severity', 'Breach Count']]
    for result in filtered_results[:50]:
        results_data.append([
            result['email'][:40] + '...' if len(result['email']) > 40 else result['email'],
            result['status'].capitalize(),
            (result.get('severity') or 'unknown').capitalize(),
            str(result.get('breach_count', 0))
        ])

    results_table = Table(results_data, colWidths=[3.0*inch, 1.1*inch, 1.2*inch, 0.9*inch])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))

    # Color code data rows by status
    for i, result in enumerate(filtered_results[:50], start=1):
        if result['status'] == 'compromised':
            bg = colors.HexColor('#FDECEC')  # light red tint
        elif result['status'] == 'clean':
            bg = colors.HexColor('#EAF7EE')  # light green tint
        elif result['status'] == 'error':
            bg = colors.HexColor('#FFF6E5')  # light yellow tint
        else:
            bg = colors.whitesmoke
        results_table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), bg)]))

    story.append(results_table)

    if len(filtered_results) > 50:
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Note: Showing first 50 of {len(filtered_results)} results. Export to Excel or CSV for complete data.",
            styles['Italic'] if 'Italic' in styles else styles['Normal']
        ))

    # Copy-friendly appendix: list all emails with breaches as bullet points
    story.append(PageBreak())
    story.append(Paragraph('Appendix: Breach Details (Copy-friendly)', styles['Heading2']))

    compromised = [r for r in results if r.get('breach_count', 0) > 0 and r.get('breaches')]
    if not compromised:
        story.append(Paragraph('No breaches found for any scanned email.', styles['Normal']))
    else:
        email_heading_style = ParagraphStyle('EmailHeading', parent=styles['Heading4'], textColor=brand_primary)
        bullet_style = ParagraphStyle('Bullet', parent=styles['Normal'], leftIndent=18, spaceBefore=2, spaceAfter=2)
        for r in compromised:
            story.append(Spacer(1, 6))
            story.append(Paragraph(r['email'], email_heading_style))
            bullets = []
            for b in r.get('breaches', []):
                title = b.get('Title') or b.get('Name') or 'Unknown Breach'
                date = b.get('BreachDate') or 'Unknown Date'
                domain = b.get('Domain') or ''
                # Keep bullet concise and copy-friendly
                text = f"{title} ({date})" + (f" — {domain}" if domain else '')
                bullets.append(ListItem(Paragraph(text, bullet_style)))
            if bullets:
                story.append(ListFlowable(bullets, bulletType='bullet', start='bulletchar'))

    # Build with header/footer on every page
    doc.build(story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    logger.info(f"PDF report exported to {filename}")
    return filename

def export_pdf(results, batch_id, timestamp, options):
    try:
        filename = write_pdf_file(results, batch_id, timestamp, options)
        return jsonify({'message': 'Results exported successfully', 'filename': filename, 'format': 'pdf'})
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 400

def export_zip_package(results, batch_id, timestamp, options):
    """Export comprehensive package with multiple formats"""
    zip_filename = f"breach_scan_package_{batch_id}_{timestamp}.zip"
    zip_filepath = os.path.join('exports', zip_filename)
    os.makedirs('exports', exist_ok=True)
    
    with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Generate files and add to the ZIP
        json_filename = write_json_file(results, batch_id, timestamp, options)
        zipf.write(os.path.join('exports', json_filename), json_filename)

        csv_filename = write_csv_file(results, batch_id, timestamp, options)
        zipf.write(os.path.join('exports', csv_filename), csv_filename)

        if EXCEL_AVAILABLE:
            try:
                excel_filename = write_excel_file(results, batch_id, timestamp, options)
                zipf.write(os.path.join('exports', excel_filename), excel_filename)
            except Exception as e:
                logger.warning(f"Excel export skipped: {e}")

        if PDF_AVAILABLE:
            try:
                pdf_filename = write_pdf_file(results, batch_id, timestamp, options)
                zipf.write(os.path.join('exports', pdf_filename), pdf_filename)
            except Exception as e:
                logger.warning(f"PDF export skipped: {e}")
        
        # Add README
        readme_content = f"""Email Breach Scanner Export Package
Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Batch ID: {batch_id}

This package contains the results of your email breach scan in multiple formats:

- JSON: Complete data with metadata and statistics
- CSV: Spreadsheet-compatible format
- Excel: Formatted spreadsheet with multiple sheets (if available)
- PDF: Professional report format (if available)

For technical support, please refer to the application documentation.
"""
        zipf.writestr("README.txt", readme_content)
    
    logger.info(f"ZIP package exported to {zip_filename}")
    return jsonify({'message': 'Package exported successfully', 'filename': zip_filename, 'format': 'zip'})

@app.route('/download/<filename>')
def download_file(filename):
    """Secure file download endpoint"""
    safe_filename = secure_filename(filename)
    filepath = os.path.join('exports', safe_filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    if not os.path.abspath(filepath).startswith(os.path.abspath('exports')):
        return jsonify({'error': 'Access denied'}), 403
    
    # Allowlist extensions produced by the app
    allowed_exts = {'.json', '.csv', '.xlsx', '.pdf', '.zip', '.txt'}
    _, ext = os.path.splitext(safe_filename)
    if ext.lower() not in allowed_exts:
        return jsonify({'error': 'File type not allowed'}), 403
    
    try:
        return send_file(filepath, as_attachment=True, download_name=safe_filename)
    except Exception as e:
        logger.error(f"Download failed for {filename}: {str(e)}")
        return jsonify({'error': 'Download failed'}), 500

@app.route('/export-formats')
def get_export_formats():
    """Get available export formats"""
    formats = {
        'json': {'name': 'JSON', 'description': 'Complete data with metadata', 'available': True},
        'csv': {'name': 'CSV', 'description': 'Spreadsheet compatible', 'available': True},
        'excel': {'name': 'Excel', 'description': 'Formatted spreadsheet', 'available': EXCEL_AVAILABLE},
        'pdf': {'name': 'PDF Report', 'description': 'Professional report', 'available': PDF_AVAILABLE},
        'zip': {'name': 'Complete Package', 'description': 'All formats in one package', 'available': True}
    }
    
    return jsonify(formats)

@app.route('/export-history')
def get_export_history():
    """Get list of exported files"""
    exports_dir = 'exports'
    if not os.path.exists(exports_dir):
        return jsonify([])
    
    files = []
    for filename in os.listdir(exports_dir):
        filepath = os.path.join(exports_dir, filename)
        if os.path.isfile(filepath):
            stat = os.stat(filepath)
            files.append({
                'filename': filename,
                'size': stat.st_size,
                'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
    
    files.sort(key=lambda x: x['created'], reverse=True)
    return jsonify(files)

if __name__ == "__main__":
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('exports', exist_ok=True)
    
    # Disable reloader to prevent background threads from duplicating
    socketio.run(
        app,
        debug=Config.DEBUG,
        host=getattr(Config, 'HOST', '0.0.0.0'),
        port=getattr(Config, 'PORT', 8000),
        use_reloader=False,
    )
